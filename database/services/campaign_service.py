import json
import re
import os
import openpyxl
import logging
import unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.schema import get_connection
from database.services.blacklist_service import is_blacklisted

_log = logging.getLogger("zapmanager.campaign")

VALID_DDDS = {
    11, 12, 13, 14, 15, 16, 17, 18, 19,
    21, 22, 24, 27, 28,
    31, 32, 33, 34, 35, 37, 38,
    41, 42, 43, 44, 45, 46, 47, 48, 49,
    51, 53, 54, 55,
    61, 62, 63, 64, 65, 66, 67, 68, 69,
    71, 73, 74, 75, 77, 79,
    81, 82, 83, 84, 85, 86, 87, 88, 89,
    91, 92, 93, 94, 95, 96, 97, 98, 99,
}

def normalize_phone(phone: str) -> str | None:
    """
    Normaliza o número de telefone para o formato E.164.
    Foca em Brasil (55), mas permite números internacionais.
    """
    try:
        if not phone: return None
        phone_str = str(phone).strip()
        digits = re.sub(r'\D', '', phone_str)
        
        if not digits: return None
        
        # Remove zeros à esquerda
        while digits.startswith('0'):
            digits = digits[1:]
            
        if not digits: return None
        
        # Caso 1: Provável número brasileiro sem DDI (10 ou 11 dígitos)
        if len(digits) in (10, 11):
            # Se não começa com 55, checamos se os 2 primeiros são um DDD válido
            ddd_cand = int(digits[:2])
            if ddd_cand in VALID_DDDS:
                # Validar se é celular (11 dígitos, 3º deve ser 9) ou fixo (10 dígitos)
                if len(digits) == 11 and digits[2] != '9':
                    # Pode ser que o 55 já estivesse lá e o DDD seja os próximos 2?
                    # Ex: 55119... (13 dígitos). Tratado no Caso 2.
                    pass 
                else:
                    return '55' + digits

        # Caso 2: Número já com DDI (incluindo 55)
        # Formato E.164 tem de 7 a 15 dígitos
        if 10 <= len(digits) <= 15:
            if digits.startswith('55'):
                # Validação extra para Brasil se for 12 ou 13 dígitos
                if len(digits) in (12, 13):
                    ddd = int(digits[2:4])
                    if ddd in VALID_DDDS:
                        return digits
                else:
                    # 55 com tamanho atípico, mas ainda pode ser válido para o WhatsApp
                    return digits
            else:
                # Número internacional (não Brasil)
                return digits
                
    except Exception:
        _log.exception("Erro ao normalizar telefone", extra={"phone": phone})
        return None

def _norm_header(header) -> str:
    if not header:
        return ""
    text = unicodedata.normalize("NFKD", str(header).lower())
    text = text.encode("ASCII", "ignore").decode()
    return text.strip()

def update_campaign_message(campaign_id: int, message_template: str) -> None:
    try:
        with get_connection() as conn:
            conn.execute("UPDATE campaigns SET message_template = ? WHERE id = ?", (message_template, campaign_id))
    except Exception:
        _log.exception("Falha ao atualizar mensagem da campanha", extra={"context": "update_campaign_message"})

def create_campaign(name: str, message_template: str, attachment_path: str = None, account_id: int = None) -> int:
    conn = get_connection()
    try:
        with conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO campaigns (name, message_template, attachment_path, account_id)
                VALUES (?, ?, ?, ?)
            ''', (name, message_template, attachment_path, account_id))
            return cursor.lastrowid
    except Exception:
        _log.exception("Falha ao criar campanha", extra={"context": "create_campaign"})
        return -1
    finally:
        conn.close()

def import_contacts_from_xlsx(campaign_id: int, xlsx_path: str) -> dict:
    PHONE_KEYS = {'numero', 'whatsapp', 'telefone', 'celular'}
    NAME_KEYS = {'nome', 'cliente', 'contato'}
    COMPANY_KEYS = {'empresa', 'razao_social', 'razao social'}
    results = {"total": 0, "imported": 0, "skipped_blacklist": 0, "duplicates_skipped": 0, "errors": []}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sheet = wb.active
        headers = [_norm_header(cell.value) for cell in sheet[1]]
        phone_col = -1
        name_col = -1
        company_col = -1
        extra_cols = []
        for i, header in enumerate(headers):
            if header in PHONE_KEYS and phone_col == -1:
                phone_col = i
            elif header in NAME_KEYS and name_col == -1:
                name_col = i
            elif header in COMPANY_KEYS and company_col == -1:
                company_col = i
            elif header:
                extra_cols.append((i, header))
        
        if phone_col == -1:
            results["errors"].append("Coluna de telefone não encontrada.")
            return results
        
        conn = get_connection()
        try:
            rows_to_insert = []
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row): continue
                results["total"] += 1
                phone_raw = row[phone_col] if phone_col < len(row) else None
                phone = normalize_phone(str(phone_raw))
                if not phone:
                    results["errors"].append(f"Número inválido na linha {index}: {phone_raw}")
                    continue
                name = str(row[name_col]).strip() if name_col != -1 and name_col < len(row) and row[name_col] else ""
                company = str(row[company_col]).strip() if company_col != -1 and company_col < len(row) and row[company_col] else ""
                extras_dict = {}
                for col_index, header in extra_cols:
                    value = row[col_index] if col_index < len(row) else None
                    if value is not None and str(value).strip():
                        extras_dict[header] = str(value).strip()
                extras_json = json.dumps(extras_dict, ensure_ascii=False)
                if is_blacklisted(phone):
                    results["skipped_blacklist"] += 1
                    conn.execute("""
                        INSERT OR IGNORE INTO campaign_contacts
                        (campaign_id, name, phone, company, extra_fields, status)
                        VALUES (?, ?, ?, ?, ?, 'BLACKLIST')
                    """, (campaign_id, name, phone, company, extras_json))
                    continue
                rows_to_insert.append((campaign_id, name, phone, company, extras_json))

            with conn:
                cursor = conn.cursor()
                cursor.executemany("INSERT OR IGNORE INTO campaign_contacts (campaign_id, name, phone, company, extra_fields) VALUES (?, ?, ?, ?, ?)", rows_to_insert)
                inserted = cursor.rowcount
                results["imported"] = inserted
                results["duplicates_skipped"] = max(0, len(rows_to_insert) - inserted)
        finally:
            conn.close()
    except Exception:
        _log.exception("Falha ao importar planilha", extra={"context": "import_contacts_from_xlsx"})
    return results

def get_pending_contacts(campaign_id: int, limit: int = None) -> list:
    conn = get_connection()
    try:
        query = "SELECT * FROM campaign_contacts WHERE campaign_id = ? AND status = 'PENDENTE' ORDER BY id ASC"
        params = [campaign_id]
        if limit:
            query += " LIMIT ?"
            params.append(limit)
        return [dict(row) for row in conn.execute(query, params).fetchall()]
    except Exception:
        _log.exception("Falha ao buscar contatos pendentes", extra={"context": "get_pending_contacts"})
        return []
    finally:
        conn.close()

def update_contact_status(contact_id: int, status: str, error_message: str = None) -> None:
    conn = get_connection()
    try:
        with conn:
            query = "UPDATE campaign_contacts SET status = ?"
            params = [status]
            if status == "ENVIADO": query += ", sent_at = datetime('now','localtime')"
            if error_message:
                query += ", error_message = ?"
                params.append(error_message)
            query += " WHERE id = ?"
            params.append(contact_id)
            conn.execute(query, params)
    except Exception:
        _log.exception("Falha ao atualizar status do contato", extra={"context": "update_contact_status"})
    finally:
        conn.close()

def get_campaign_stats(campaign_id: int) -> dict:
    conn = get_connection()
    try:
        row = conn.execute("""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status='ENVIADO' THEN 1 ELSE 0 END) AS sent,
                SUM(CASE WHEN status='ERRO' THEN 1 ELSE 0 END) AS failed,
                SUM(CASE WHEN status='INVÁLIDO' THEN 1 ELSE 0 END) AS invalid,
                SUM(CASE WHEN status='PENDENTE' THEN 1 ELSE 0 END) AS pending,
                SUM(CASE WHEN status='BLACKLIST' THEN 1 ELSE 0 END) AS blacklist
            FROM campaign_contacts WHERE campaign_id = ?
        """, (campaign_id,)).fetchone()
        res = dict(row) if row else {"total": 0, "sent": 0, "failed": 0, "invalid": 0, "pending": 0, "blacklist": 0}
        return {k: (v if v is not None else 0) for k, v in res.items()}
    except Exception:
        _log.exception("Falha ao obter estatísticas da campanha", extra={"context": "get_campaign_stats"})
        return {"total": 0, "sent": 0, "failed": 0, "invalid": 0, "pending": 0, "blacklist": 0}
    finally:
        conn.close()

def get_campaign_history() -> list:
    conn = get_connection()
    try:
        return [dict(row) for row in conn.execute("SELECT * FROM campaigns ORDER BY created_at DESC").fetchall()]
    except Exception:
        _log.exception("Falha ao ler histórico de campanhas", extra={"context": "get_campaign_history"})
        return []
    finally:
        conn.close()

def reset_failed_contacts(campaign_id: int) -> int:
    conn = get_connection()
    try:
        with conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE campaign_contacts SET status = 'PENDENTE', error_message = NULL WHERE campaign_id = ? AND status = 'ERRO'", (campaign_id,))
            return cursor.rowcount
    except Exception:
        _log.exception("Falha ao resetar contatos com erro", extra={"context": "reset_failed_contacts"})
        return 0
    finally:
        conn.close()

def get_campaign_details(campaign_id: int) -> dict:
    conn = get_connection()
    try:
        campaign = conn.execute("SELECT * FROM campaigns WHERE id = ?", (campaign_id,)).fetchone()
        if not campaign: return None
        contacts = conn.execute("SELECT * FROM campaign_contacts WHERE campaign_id = ? ORDER BY id ASC", (campaign_id,)).fetchall()
        return {"campaign": dict(campaign), "contacts": [dict(c) for c in contacts]}
    except Exception:
        _log.exception("Falha ao buscar detalhes da campanha", extra={"context": "get_campaign_details"})
        return None
    finally:
        conn.close()

def reset_stuck_contacts():
    conn = get_connection()
    try:
        with conn:
            conn.execute("UPDATE campaign_contacts SET status = 'PENDENTE', error_message = NULL WHERE status = 'EM_PROCESSAMENTO'")
    except Exception:
        _log.exception("Falha ao resetar contatos presos", extra={"context": "reset_stuck_contacts"})
    finally:
        conn.close()

def reset_processing_contacts(campaign_id: int) -> int:
    conn = get_connection()
    try:
        with conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE campaign_contacts SET status = 'PENDENTE' WHERE campaign_id = ? AND status = 'EM_PROCESSAMENTO'", (campaign_id,))
            return cursor.rowcount
    except Exception:
        _log.exception("Falha ao resetar contatos em processamento", extra={"context": "reset_processing_contacts"})
        return 0
    finally:
        conn.close()

def get_contact_attachment(contact_id: int):
    conn = get_connection()
    try:
        row = conn.execute("SELECT attachment_path FROM campaign_contacts WHERE id = ?", (contact_id,)).fetchone()
        return row["attachment_path"] if row else None
    except Exception:
        _log.exception("Falha ao buscar anexo do contato", extra={"context": "get_contact_attachment"})
        return None
    finally:
        conn.close()

def update_contact_attachment(contact_id: int, attachment_path) -> bool:
    conn = get_connection()
    try:
        with conn:
            conn.execute("UPDATE campaign_contacts SET attachment_path = ? WHERE id = ?", (attachment_path, contact_id))
        return True
    except Exception:
        _log.exception("Falha ao atualizar anexo do contato", extra={"context": "update_contact_attachment"})
        return False
    finally:
        conn.close()

def export_campaign_to_xlsx(campaign_id: int, output_path: str) -> str | None:
    """Gera um arquivo Excel com o resultado dos disparos da campanha."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        details = get_campaign_details(campaign_id)
        if not details:
            return None
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Campanha"
        
        # Cabeçalho
        headers = ["Nome", "Telefone", "Empresa", "Status", "Erro / Observação"]
        ws.append(headers)
        
        # Estilo cabeçalho
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Dados
        for c in details["contacts"]:
            ws.append([
                c.get("name", ""),
                c.get("phone", ""),
                c.get("company", ""),
                c.get("status", ""),
                c.get("error_message") or c.get("observation") or ""
            ])
            
        # Ajustar largura colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
            
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return output_path
    except Exception:
        _log.exception("Falha ao exportar campanha para Excel", extra={"context": "export_campaign_to_xlsx", "campaign_id": campaign_id})
        return None

def get_today_sent_count() -> int:
    """
    Retorna o total de contatos marcados como ENVIADO hoje (por sent_at)
    em todas as campanhas. Usado para aplicar limite diario por plano.
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM campaign_contacts
            WHERE status = 'ENVIADO'
            AND DATE(sent_at) = DATE('now', 'localtime')
        """)
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else 0
    except Exception:
        _log.exception("Falha ao contar envios do dia", extra={"context": "get_today_sent_count"})
        return 0
