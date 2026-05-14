import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_template_workbook() -> openpyxl.Workbook:
    """Monta a planilha modelo de contatos em memoria."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatos"

    headers = ["Nome", "Numero", "Empresa", "Adicional1", "Adicional2", "Adicional3"]
    header_fill = PatternFill("solid", fgColor="00A884")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    rows = [
        ("Maria Silva", "11999998888", "Empresa Alpha", "Promo50", "ABC123", "Black Friday"),
        ("Joao Souza", "21988887777", "Beta Ltda", "Promo30", "DEF456", "Cyber Monday"),
        ("Ana Paula", "31977776666", "", "Promo20", "GHI789", ""),
        ("Carlos Mendes", "11133334444", "Gamma S/A", "", "", ""),
        ("Teste Invalido", "000", "", "", "", ""),
    ]

    alt_fill = PatternFill("solid", fgColor="F0FAF7")
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = alt_fill

    widths = [24, 16, 22, 18, 18, 22]
    for col_index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return wb
