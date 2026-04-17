import os
import time
import random
import json
import re
import shutil
import logging
import threading
import urllib.parse
import queue
from datetime import datetime
import customtkinter as ctk
from tkinter import *
from tkinter import filedialog, messagebox, ttk

# Bibliotecas de Automação
import openpyxl
import requests
import json
import subprocess
import signal
import socket

# ===============================================================
# CONFIGURAÇÃO DE LOGS EM ARQUIVO
# ===============================================================
def _setup_file_logger():
    if not os.path.exists("logs"):
        os.makedirs("logs")
    log_file = os.path.join("logs", f"zap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format="%(asctime)s [%(levelname)s] %(message)s",
        encoding="utf-8"
    )
    return log_file

# ===============================================================
# SERVIÇO DE CONFIGURAÇÃO (PERSISTÊNCIA)
# ===============================================================
class ConfigService:
    FILE = "zap_config.json"

    @staticmethod
    def load():
        if os.path.exists(ConfigService.FILE):
            try:
                with open(ConfigService.FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError):
                pass
        return {}

    @staticmethod
    def save(config):
        try:
            with open(ConfigService.FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
        except IOError as e:
            logging.error(f"Falha ao salvar configuração: {e}")

# ===============================================================
# SERVIÇO DE EXCEL E BACKUP
# ===============================================================
class ExcelService:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active
        self.cols = self._smart_map()

    def _smart_map(self):
        header = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in self.sheet[1]]
        mappings = {
            'nome':       [r'nome', r'cliente', r'contato'],
            'numero':     [r'whatsapp', r'n[uú]mero', r'telefone', r'celular', r'phone'],
            'status':     [r'status', r'situa[cç][aã]o'],
            'empresa':    [r'empresa', r'raz[aã]o social', r'fantasia'],
            'observacao': [r'observa[cç][aã]o', r'^obs'],
            'dataenvio':  [r'dataenvio', r'data de envio', r'enviado em']
        }
        mapped = {}
        for key, patterns in mappings.items():
            for i, h in enumerate(header):
                if any(re.search(p, h) for p in patterns):
                    mapped[key] = i + 1
                    break

        missing = [r for r in ['nome', 'numero', 'status'] if r not in mapped]
        if missing:
            raise ValueError(
                f"Coluna(s) obrigatória(s) não encontrada(s): {', '.join(m.upper() for m in missing)}. "
                "Verifique o cabeçalho da planilha."
            )

        if 'observacao' not in mapped:
            mapped['observacao'] = self._add_col("Observacao")
        if 'dataenvio' not in mapped:
            mapped['dataenvio'] = self._add_col("DataEnvio")
        return mapped

    def _add_col(self, name):
        col = self.sheet.max_column + 1
        self.sheet.cell(row=1, column=col).value = name
        self.workbook.save(self.file_path)
        return col

    def create_backup(self):
        if not os.path.exists("backups"):
            os.makedirs("backups")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join("backups", f"BKP_{ts}_{os.path.basename(self.file_path)}")
        shutil.copy2(self.file_path, dest)
        logging.info(f"Backup criado: {dest}")
        return dest

    def recover_stuck_rows(self):
        recovered = 0
        for row in range(2, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=self.cols['status'])
            if str(cell.value).strip().upper() == "EM_PROCESSAMENTO":
                cell.value = "PENDENTE"
                recovered += 1
        if recovered > 0:
            self.workbook.save(self.file_path)
            logging.info(f"{recovered} linha(s) recuperada(s) de EM_PROCESSAMENTO para PENDENTE.")
        return recovered

    def count_pending(self):
        count = 0
        pending_values = {'PENDENTE', 'NONE', '', 'NAN'}
        for row in range(2, self.sheet.max_row + 1):
            val = str(self.sheet.cell(row=row, column=self.cols['status']).value).strip().upper()
            if val in pending_values:
                count += 1
        return count

    def update_row(self, row_idx, status, obs="", sent=False):
        self.sheet.cell(row=row_idx, column=self.cols['status']).value = status
        if obs:
            self.sheet.cell(row=row_idx, column=self.cols['observacao']).value = obs
        if sent:
            self.sheet.cell(row=row_idx, column=self.cols['dataenvio']).value = \
                datetime.now().strftime("%d/%m/%Y %H:%M")
        try:
            self.workbook.save(self.file_path)
        except PermissionError:
            logging.error(
                f"PermissionError ao salvar a planilha (linha {row_idx}). "
                "Feche o arquivo no Excel e tente novamente."
            )
            raise PermissionError(
                "Não foi possível salvar a planilha. Feche o arquivo no Microsoft Excel e tente novamente."
            )

# ===============================================================
# MOTOR DE AUTOMAÇÃO
# ===============================================================
class AutomationEngine:
    def __init__(self, log_queue):
        self.log_queue = log_queue
        self.base_url = "http://localhost:3001"
        self.process = None
        self.motor_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "whatsapp-motor")

    def log(self, msg, level="INFO"):
        self.log_queue.put((msg, level))
        if level == "ERROR":
            logging.error(msg)
        elif level == "WARN":
            logging.warning(msg)
        else:
            logging.info(msg)

    def _is_server_running(self):
        try:
            with socket.create_connection(("127.0.0.1", 3001), timeout=1):
                return True
        except (socket.timeout, ConnectionRefusedError):
            return False

    def start(self):
        """Inicia o Motor Node.js e aguarda a conexão do WhatsApp."""
        self.log("Verificando Motor de Automação (Node.js)...", "INFO")
        
        # Se já estiver rodando, só validamos o status
        if not self._is_server_running():
            self.log("Iniciando Motor Node.js em segundo plano...", "INFO")
            
            node_exe = "node"
            # Tenta localizar node no caminho padrão se necessário
            if not self._check_node_in_path():
                node_exe = r"C:\Program Files\nodejs\node.exe"
                if not os.path.exists(node_exe):
                    self.log("Node.js não localizado! Por favor instale de nodejs.org", "ERROR")
                    return False

            try:
                # No Windows, usamos CREATE_NO_WINDOW para rodar invisível
                flags = 0
                if os.name == 'nt':
                    # CREATE_NO_WINDOW = 0x08000000
                    flags = 0x08000000
                
                self.process = subprocess.Popen(
                    [node_exe, "server.js"],
                    cwd=self.motor_dir,
                    creationflags=flags
                )
                
                # Nota: Como não estamos capturando o stdout (para que o usuário veja o QR na janela),
                # o monitoramento de logs via _monitor_output é desativado para evitar bloqueio.
                
            except Exception as e:
                self.log(f"Falha ao executar Motor Node: {e}", "ERROR")
                return False

        # Aguarda o servidor HTTP subir
        self.log("Aguardando inicialização do servidor local...", "INFO")
        for _ in range(30):
            if self._is_server_running():
                break
            time.sleep(1)
        else:
            self.log("Erro: Servidor Node.js não respondeu na porta 3001.", "ERROR")
            return False

        # Loop de espera para o WhatsApp estar "READY" (autenticado)
        self.log("Aguardando WhatsApp estar pronto (READY/LOGIN)...", "INFO")
        self.log("Dica: Se for a primeira vez, verifique o console do Node para o QR Code.", "WARN")
        
        start_time = time.time()
        while time.time() - start_time < 300:  # Timeout 5 minutos
            try:
                resp = requests.get(f"{self.base_url}/status", timeout=5).json()
                if resp.get("connected"):
                    self.log("✅ Conexão com WhatsApp estabelecida!", "SUCCESS")
                    return True
            except:
                pass
            time.sleep(3)
        
        self.log("Timeout aguardando login do WhatsApp.", "ERROR")
        return False


    def _check_node_in_path(self):
        import shutil
        return shutil.which("node") is not None

    def _monitor_output(self):
        """Lê a saída do Node.js e repassa para a fila de logs (útil para ver o QR Code se necessário)."""
        if not self.process: return
        for line in iter(self.process.stdout.readline, ''):
            clean_line = line.strip()
            if clean_line:
                if "SCANEAR QR CODE" in clean_line:
                    # Notifica o usuário que precisa do celular
                    self.log("⚠️ LOGIN NECESSÁRIO: Scaneie o QR Code no terminal do motor.", "WARN")
                elif "PRONTO PARA USO" in clean_line:
                    self.log("Motor sinalizou prontidão.", "INFO")
                # logs comuns do Node vão para o console do Python
                print(f"[NodeJS] {clean_line}")

    def send_message(self, number, message):
        """Envia apenas texto via API Node."""
        try:
            payload = {
                "number": number,
                "message": message
            }
            resp = requests.post(f"{self.base_url}/send", json=payload, timeout=30)
            if resp.status_code == 200:
                return "SUCESSO"
            
            err = resp.json().get("error", "Erro desconhecido")
            if "invalid" in err.lower():
                return "INVALIDO"
            
            self.log(f"Erro no envio: {err}", "ERROR")
            return "ERRO"
        except Exception as e:
            self.log(f"Exceção no envio: {e}", "ERROR")
            return "ERRO"

    def send_with_attachment(self, number, message, attachment_path):
        """Envia arquivo (foto/doc) com legenda unificada via API Node."""
        try:
            payload = {
                "number": number,
                "message": message,
                "filePath": os.path.abspath(attachment_path)
            }
            # O Motor Node.js já trata se é foto ou documento internamente.
            resp = requests.post(f"{self.base_url}/send", json=payload, timeout=60)
            
            if resp.status_code == 200:
                return "SUCESSO"
            
            err = resp.json().get("error", "Erro desconhecido")
            self.log(f"Erro no anexo: {err}", "ERROR")
            if "invalid" in err.lower():
                return "INVALIDO"
            return "ERRO"
        except Exception as e:
            self.log(f"Exceção no anexo: {e}", "ERROR")
            return "ERRO"

    def stop(self):
        """Finaliza o motor se necessário."""
        if self.process:
            try:
                self.process.terminate()
                self.process = None
                self.log("Motor desativado.", "INFO")
            except:
                pass

# ===============================================================
# INTERFACE E ORQUESTRAÇÃO
# ===============================================================
class ZapAutomationApp:
    def __init__(self, root):
        self.root = root
        
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")
        
        self.root.title("ZapManager Pro v4.5 (Dashboard Apple)")
        self.root.geometry("1050x750")
        
        self.log_queue = queue.Queue()
        self.running = False
        self.stop_req = False
        self.excel = None
        self.engine = None
        self.current_log_file = None

        self._setup_ui()
        self._load_config()
        self.root.after(100, self._process_logs)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _setup_ui(self):
        # Configure grid for standard Dashboard (Sidebar + Main Content)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        # ─── SIDEBAR (Esquerda) ───
        sidebar_frame = ctk.CTkFrame(self.root, width=240, corner_radius=0)
        sidebar_frame.grid(row=0, column=0, sticky="nsew")
        sidebar_frame.grid_rowconfigure(5, weight=1)

        logo_label = ctk.CTkLabel(sidebar_frame, text="ZapManager Pro", font=ctk.CTkFont(size=22, weight="bold"))
        logo_label.grid(row=0, column=0, padx=20, pady=(30, 10))
        
        self.status_badge = ctk.CTkLabel(sidebar_frame, text="Aguardando Planilha", fg_color="#1A1A1A", text_color="#FFFFFF", corner_radius=5)
        self.status_badge.grid(row=1, column=0, padx=20, pady=5, sticky="ew")

        self.lbl_pending = ctk.CTkLabel(sidebar_frame, text="0 pendente(s)", font=ctk.CTkFont(size=14))
        self.lbl_pending.grid(row=2, column=0, padx=20, pady=(20, 5))
        
        self.progress_bar = ctk.CTkProgressBar(sidebar_frame, progress_color="#FF6600")
        self.progress_bar.grid(row=3, column=0, padx=20, pady=(0, 20), sticky="ew")
        self.progress_bar.set(0)

        self.btn_run = ctk.CTkButton(
            sidebar_frame, text="▶ INICIAR", command=self._start,
            fg_color="#FF6600", hover_color="#FF5011", font=ctk.CTkFont(weight="bold", size=14), height=40
        )
        self.btn_run.grid(row=4, column=0, padx=20, pady=10, sticky="ew")

        self.btn_stop = ctk.CTkButton(
            sidebar_frame, text="■ PARAR", command=self._stop,
            fg_color="#1A1A1A", hover_color="#0F172A", font=ctk.CTkFont(weight="bold", size=14), height=40, state="disabled"
        )
        self.btn_stop.grid(row=6, column=0, padx=20, pady=20, sticky="ew")

        # ─── MAIN CONTENT (Direita) ───
        main_view = ctk.CTkScrollableFrame(self.root, fg_color="transparent")
        main_view.grid(row=0, column=1, sticky="nsew", padx=20, pady=0)
        main_view.grid_columnconfigure(0, weight=1)

        # 1. Card Planilha
        card_excel = ctk.CTkFrame(main_view, corner_radius=10)
        card_excel.grid(row=0, column=0, padx=10, pady=(20, 10), sticky="ew")
        card_excel.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(card_excel, text="1. Fonte de Dados", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")
        
        file_frame = ctk.CTkFrame(card_excel, fg_color="transparent")
        file_frame.grid(row=1, column=0, padx=20, pady=(0, 15), sticky="ew")
        self.btn_file = ctk.CTkButton(file_frame, text="Selecionar Planilha...", command=self._load_file_dialog, width=200, fg_color="#00CCFF", hover_color="#00E5E5", text_color="#1A1A1A")
        self.btn_file.pack(side="left")
        self.lbl_file = ctk.CTkLabel(file_frame, text="Nenhum arquivo selecionado.", text_color="#9CA3AF")
        self.lbl_file.pack(side="left", padx=15)

        # 2. Card Mensagem e Anexo
        card_msg = ctk.CTkFrame(main_view, corner_radius=10)
        card_msg.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        card_msg.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(card_msg, text="2. Conteúdo da Mensagem", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")
        ctk.CTkLabel(card_msg, text="Use {nome} e {empresa} para personalizar a mensagem.", text_color="#9CA3AF").grid(row=1, column=0, padx=20, pady=0, sticky="w")
        
        self.txt_msg = ctk.CTkTextbox(card_msg, height=100, corner_radius=8, border_width=1)
        self.txt_msg.grid(row=2, column=0, padx=20, pady=(10, 15), sticky="ew")

        att_frame = ctk.CTkFrame(card_msg, fg_color="transparent")
        att_frame.grid(row=3, column=0, padx=20, pady=(0, 15), sticky="ew")
        
        self.var_attachment = ctk.BooleanVar(value=False)
        self.chk_attachment = ctk.CTkSwitch(att_frame, text="Incluir Anexo", variable=self.var_attachment, command=self._toggle_attachment, progress_color="#FF6600")
        self.chk_attachment.pack(side="left")

        self.btn_attachment = ctk.CTkButton(att_frame, text="Escolher Arquivo...", command=self._select_attachment, fg_color="#FF7A1A", hover_color="#FF5011", text_color="#FFFFFF", state="disabled", width=140)
        self.btn_attachment.pack(side="left", padx=(20, 10))

        self.lbl_attachment = ctk.CTkLabel(att_frame, text="", text_color="#9CA3AF")
        self.lbl_attachment.pack(side="left")
        
        self.btn_clear_att = ctk.CTkButton(att_frame, text="✕", command=self._clear_attachment, fg_color="#4B5563", hover_color="#1A1A1A", width=30, state="disabled")
        self.btn_clear_att.pack(side="left", padx=5)
        self.attachment_path = ""

        # 3. Card Configurações
        card_cfg = ctk.CTkFrame(main_view, corner_radius=10)
        card_cfg.grid(row=2, column=0, padx=10, pady=10, sticky="ew")
        card_cfg.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(card_cfg, text="3. Regras de Disparo", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        cfg_r1 = ctk.CTkFrame(card_cfg, fg_color="transparent")
        cfg_r1.grid(row=1, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="ew")
        
        ctk.CTkLabel(cfg_r1, text="Parar após (qtd):").pack(side="left")
        self.ent_limit = ctk.CTkEntry(cfg_r1, width=60)
        self.ent_limit.pack(side="left", padx=(10, 30))

        self.var_keep_open = ctk.BooleanVar(value=False)
        self.chk_keep = ctk.CTkSwitch(cfg_r1, text="Manter navegador aberto ao final", variable=self.var_keep_open, progress_color="#FF6600")
        self.chk_keep.pack(side="left")

        cfg_r2 = ctk.CTkFrame(card_cfg, fg_color="transparent")
        cfg_r2.grid(row=2, column=0, columnspan=2, padx=20, pady=(0, 15), sticky="ew")

        ctk.CTkLabel(cfg_r2, text="Intervalo aleatório:").pack(side="left")
        self.ent_min = ctk.CTkEntry(cfg_r2, width=50)
        self.ent_min.pack(side="left", padx=(10, 5))
        ctk.CTkLabel(cfg_r2, text="até").pack(side="left")
        self.ent_max = ctk.CTkEntry(cfg_r2, width=50)
        self.ent_max.pack(side="left", padx=5)
        ctk.CTkLabel(cfg_r2, text="segundos").pack(side="left")

        # 4. Terminal de Logs Embutido
        self.log_frame = ctk.CTkFrame(main_view, corner_radius=10)
        self.log_frame.grid(row=3, column=0, padx=10, pady=(10, 20), sticky="ew")
        self.log_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.log_frame, text="Console de Atividades", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, padx=15, pady=(10, 5), sticky="w")
        
        self.log_area = Text(
            self.log_frame, height=12, bg="#0F172A", fg="#00CCFF",
            font=("Consolas", 10), state=DISABLED, relief=FLAT, padx=10, pady=10
        )
        self.log_area.grid(row=1, column=0, padx=15, pady=(0, 15), sticky="ew")

    # ── Helpers de UI ──────────────────────────────────────────

    def _toggle_attachment(self):
        """Ativa/desativa o botão de seleção de arquivo."""
        if self.var_attachment.get():
            self.btn_attachment.configure(state="normal")
        else:
            self.btn_attachment.configure(state="disabled")
            self.btn_clear_att.configure(state="disabled")
            self._clear_attachment()

    def _select_attachment(self):
        extensions = [
            ("Imagens",    "*.jpg *.jpeg *.png *.gif *.webp *.bmp"),
            ("Vídeos",     "*.mp4 *.3gp *.mov *.avi"),
            ("Documentos", "*.pdf *.doc *.docx *.xls *.xlsx *.ppt *.pptx *.txt *.zip"),
            ("Todos",      "*.*"),
        ]
        path = filedialog.askopenfilename(filetypes=extensions)
        if path:
            self.attachment_path = path
            name = os.path.basename(path)
            display = name if len(name) <= 35 else name[:32] + "..."
            self.lbl_attachment.configure(text=display, text_color=("#111827", "#f9fafb"))
            self.btn_clear_att.configure(state="normal")
            self._set_status("Anexo configurado")

    def _clear_attachment(self):
        self.attachment_path = ""
        self.lbl_attachment.configure(text="", text_color="#9CA3AF")
        self.btn_clear_att.configure(state="disabled")
        self._set_status("Anexo removido")

    def _set_status(self, msg):
        self.status_badge.configure(text=msg)
        if "andamento" in msg.lower():
            self.status_badge.configure(fg_color="#FF6600", text_color="#FFFFFF")
        elif "Pronto" in msg:
            self.status_badge.configure(fg_color="#1A1A1A", text_color="#FFFFFF")
        else:
            self.status_badge.configure(fg_color="#00CCFF", text_color="#1A1A1A")

    def _update_pending_label(self):
        if self.excel:
            count = self.excel.count_pending()
            self.lbl_pending.configure(text=f"{count} pendente(s)")

    # ── Carregamento de arquivo ────────────────────────────────

    def _load_file_dialog(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self._load_file(path)

    def _load_file(self, path):
        try:
            self.excel = ExcelService(path)
            self.lbl_file.configure(text=os.path.basename(path), text_color=("#111827", "#f9fafb"))
            recovered = self.excel.recover_stuck_rows()
            if recovered:
                self.log_queue.put((f"{recovered} contato(s) recuperado(s) de EM_PROCESSAMENTO → PENDENTE.", "WARN"))
            self._update_pending_label()
            self._set_status("Planilha carregada")
        except PermissionError:
            messagebox.showerror(
                "Arquivo em uso",
                "A planilha está aberta em outro programa (ex: Excel).\n"
                "Feche-a e tente novamente."
            )
        except Exception as e:
            messagebox.showerror("Erro ao carregar planilha", str(e))

    # ── Configuração ───────────────────────────────────────────

    def _load_config(self):
        cfg = ConfigService.load()
        msg_saved = cfg.get("msg", "Olá {nome}! Tudo bem?")
        self.txt_msg.insert("1.0", msg_saved)
        self.ent_limit.insert(0, cfg.get("limit", "100"))
        self.ent_min.insert(0, cfg.get("min", "15"))
        self.ent_max.insert(0, cfg.get("max", "30"))
        self.var_keep_open.set(cfg.get("keep_open", False))
        last_f = cfg.get("last_file", "")
        if last_f and os.path.exists(last_f):
            self._load_file(last_f)
        # Restaurar anexo salvo anteriormente
        last_att = cfg.get("last_attachment", "")
        if last_att and os.path.exists(last_att):
            self.attachment_path = last_att
            self.var_attachment.set(True)
            self.btn_attachment.configure(state="normal")
            self.btn_clear_att.configure(state="normal")
            name = os.path.basename(last_att)
            display = name if len(name) <= 35 else name[:32] + "..."
            self.lbl_attachment.configure(text=display, text_color=("#111827", "#f9fafb"))

    def _save_config(self, params):
        ConfigService.save({
            "msg": params["msg"],
            "limit": params["limit"],
            "min": params["min"],
            "max": params["max"],
            "keep_open": self.var_keep_open.get(),
            "last_file": self.excel.file_path if self.excel else "",
            "last_attachment": params.get("attachment", "")
        })

    # ── Processamento de logs (thread-safe) ───────────────────

    def _process_logs(self):
        while not self.log_queue.empty():
            msg, level = self.log_queue.get()
            tag_color = {
                "ERROR": "#FF5011",
                "WARN": "#FF7A1A",
                "SUCCESS": "#00CCFF",
                "MSG_BOX": "#FFFFFF",
            }.get(level, "#9CA3AF")

            self.log_area.config(state=NORMAL)
            self.log_area.insert(
                END,
                f"[{datetime.now().strftime('%H:%M:%S')}] {level}: {msg}\n",
                level
            )
            self.log_area.tag_config(level, foreground=tag_color)
            self.log_area.see(END)
            self.log_area.config(state=DISABLED)

            if level == "MSG_BOX":
                messagebox.showinfo("ZapManager Pro", msg)

        self.root.after(150, self._process_logs)

    # ── Controle de estado da UI ───────────────────────────────

    def _set_ui_state(self, state):
        s_str = "normal" if state == NORMAL else "disabled"
        for w in [self.btn_file, self.txt_msg, self.ent_limit,
                  self.ent_min, self.ent_max]:
            try:
                w.configure(state=s_str)
            except Exception:
                pass
        
        if state == NORMAL:
            self.btn_run.configure(state="normal")
            self.btn_stop.configure(state="disabled")
        else:
            self.btn_run.configure(state="disabled")
            self.btn_stop.configure(state="normal")

    # ── Validação de número ────────────────────────────────────

    def _validate_number(self, raw_num):
        if raw_num is None:
            return None
        num = "".join(filter(str.isdigit, str(raw_num)))
        if not num:
            return None

        # Brasil: DDD (2 dígitos) + número (8 ou 9 dígitos) → 10 ou 11 dígitos
        if 10 <= len(num) <= 11:
            num = "55" + num

        # Aceita 12 ou 13 dígitos (ex: 5511999998888)
        if len(num) < 12 or len(num) > 13:
            return None

        return num

    # ── Início/Parada de campanha ──────────────────────────────

    def _start(self):
        if not self.excel:
            return messagebox.showwarning("Atenção", "Selecione a planilha antes de iniciar.")

        try:
            msg_text = self.txt_msg.get("1.0", "end-1c").strip()
            if not msg_text:
                return messagebox.showwarning("Atenção", "A mensagem não pode estar vazia.")

            limit_val = int(self.ent_limit.get())
            min_val = int(self.ent_min.get())
            max_val = int(self.ent_max.get())

            if limit_val <= 0:
                raise ValueError("O limite deve ser maior que zero.")
            if min_val > max_val:
                raise ValueError("O intervalo mínimo não pode ser maior que o máximo.")
            if min_val < 5:
                messagebox.showwarning(
                    "Intervalo baixo",
                    "Intervalos abaixo de 5 segundos podem causar instabilidade. "
                    "Recomendamos no mínimo 15 segundos."
                )

            params = {
                "msg": msg_text,
                "limit": limit_val,
                "min": min_val,
                "max": max_val,
                "keep_open": self.var_keep_open.get(),
                "attachment": ""
            }

            # Valida anexo se ativado
            if self.var_attachment.get():
                att = self.attachment_path.strip()
                if not att:
                    return messagebox.showwarning(
                        "Anexo não selecionado",
                        "A opção de anexo está marcada, mas nenhum arquivo foi selecionado.\n"
                        "Selecione um arquivo ou desmarque a opção de anexo."
                    )
                if not os.path.exists(att):
                    return messagebox.showerror(
                        "Arquivo não encontrado",
                        f"O arquivo de anexo não existe:\n{att}\n\nSelecione o arquivo novamente."
                    )
                params["attachment"] = att

        except ValueError as e:
            return messagebox.showerror("Configuração inválida", str(e))

        self.current_log_file = _setup_file_logger()
        logging.info(f"Campanha iniciada. Limite: {params['limit']}, Intervalo: {params['min']}-{params['max']}s")

        self._save_config(params)
        self.running = True
        self.stop_req = False
        self.progress_bar["value"] = 0
        self._set_ui_state(DISABLED)
        self._set_status("Campanha em andamento...")
        threading.Thread(target=self._execute, args=(params,), daemon=True).start()

    def _stop(self):
        if self.running:
            self.stop_req = True
            self.log_queue.put(("Parada solicitada. Aguardando finalizar o contato atual...", "WARN"))
            self._set_status("Parando campanha...")

    def _on_close(self):
        if self.running:
            if not messagebox.askyesno(
                "Sair", "Uma campanha está em andamento. Deseja mesmo sair?\nO progresso atual será salvo."
            ):
                return
            self.stop_req = True
        self.root.destroy()

    # ── Thread de execução ────────────────────────────────────

    def _execute(self, params):
        self.engine = None
        processed = 0
        total_pending = self.excel.count_pending()

        try:
            bkp = self.excel.create_backup()
            self.log_queue.put((f"Backup criado: {os.path.basename(bkp)}", "INFO"))

            self.engine = AutomationEngine(self.log_queue)
            if not self.engine.start():
                self.log_queue.put(("Não foi possível iniciar o Chrome. Campanha cancelada.", "ERROR"))
                return

            for row in range(2, self.excel.sheet.max_row + 1):
                if self.stop_req or processed >= params["limit"]:
                    break

                status_val = str(
                    self.excel.sheet.cell(row=row, column=self.excel.cols['status']).value
                ).strip().upper()

                if status_val not in ('PENDENTE', 'NONE', '', 'NAN'):
                    continue

                # Marca como em processamento antes de tentar
                self.excel.update_row(row, "EM_PROCESSAMENTO")

                nome = str(
                    self.excel.sheet.cell(row=row, column=self.excel.cols['nome']).value or "Cliente"
                ).strip()
                raw_num = self.excel.sheet.cell(row=row, column=self.excel.cols['numero']).value
                empresa_col = self.excel.cols.get('empresa')
                empresa = str(
                    self.excel.sheet.cell(row=row, column=empresa_col).value if empresa_col else ""
                ).strip() or ""

                num = self._validate_number(raw_num)
                if not num:
                    self.excel.update_row(row, "INVALIDO", f"Número fora do padrão: {raw_num}")
                    self.log_queue.put((f"[Linha {row}] {nome}: número inválido ({raw_num})", "WARN"))
                    continue

                msg = params["msg"].replace("{nome}", nome).replace("{empresa}", empresa)
                att = params.get("attachment", "")

                if att:
                    self.log_queue.put((f"[Linha {row}] Enviando com anexo para {nome} ({num[-4:]}...)...", "INFO"))
                else:
                    self.log_queue.put((f"[Linha {row}] Enviando para {nome} ({num[-4:]}...)...", "INFO"))

                try:
                    if att:
                        res = self.engine.send_with_attachment(num, msg, att)
                    else:
                        res = self.engine.send_message(num, msg)
                except Exception as e:
                    res = "ERRO"
                    logging.error(f"Linha {row} – exceção no envio: {e}")

                if res == "SUCESSO":
                    self.excel.update_row(row, "ENVIADO", sent=True)
                    processed += 1
                    self.log_queue.put((f"✔ {nome} — ENVIADO ({processed}/{params['limit']})", "SUCCESS"))
                elif res == "INVALIDO":
                    self.excel.update_row(row, "INVALIDO", "WhatsApp informou número inválido")
                    self.log_queue.put((f"✘ {nome} — INVÁLIDO", "WARN"))
                else:
                    self.excel.update_row(row, "ERRO", "Falha técnica na entrega")
                    self.log_queue.put((f"✘ {nome} — ERRO técnico", "ERROR"))

                # Progresso
                pct = min((processed / params["limit"]) * 100, 100)
                self.root.after(0, lambda v=pct: self.progress_bar.set(v / 100.0))
                self.root.after(0, self._update_pending_label)

                # Intervalo aleatório entre envios (apenas se houver próximo)
                if not self.stop_req and processed < params["limit"]:
                    wait_time = random.randint(params["min"], params["max"])
                    self.log_queue.put((f"Aguardando {wait_time}s antes do próximo...", "INFO"))
                    time.sleep(wait_time)

            motivo = "Limite atingido" if processed >= params["limit"] else ("Parado pelo usuário" if self.stop_req else "Lista finalizada")
            summary = f"Campanha encerrada! {processed} enviado(s). Motivo: {motivo}"
            self.log_queue.put((summary, "MSG_BOX"))
            logging.info(summary)

        except PermissionError as e:
            self.log_queue.put((str(e), "ERROR"))
        except Exception as e:
            self.log_queue.put((f"Erro inesperado: {str(e)}", "ERROR"))
            logging.exception("Erro inesperado na thread de execução")
        finally:
            self.running = False
            if self.engine and not params.get("keep_open", False):
                self.engine.stop()

            self.root.after(0, lambda: self._set_ui_state(NORMAL))
            self.root.after(0, lambda: self._set_status(f"Pronto. {processed} mensagem(ns) enviada(s)."))
            self.root.after(0, self._update_pending_label)
            self.root.after(0, lambda: self.progress_bar.set(1.0 if processed > 0 else 0.0))


if __name__ == "__main__":
    ctk.set_appearance_mode("System")
    root = ctk.CTk()
    app = ZapAutomationApp(root)
    root.mainloop()
