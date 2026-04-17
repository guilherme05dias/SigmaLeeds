"""
criar_planilha_modelo.py — Execute este script UMA VEZ para gerar a planilha modelo.
Uso: python criar_planilha_modelo.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Contatos"

headers = ["Nome", "Numero", "Status", "Empresa", "Observacao", "DataEnvio"]
header_fill = PatternFill("solid", fgColor="00A884")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

rows = [
    ("Maria Silva",    "11999998888", "PENDENTE", "Empresa Alpha", "", ""),
    ("Joao Souza",     "21988887777", "PENDENTE", "Beta Ltda",     "", ""),
    ("Ana Paula",      "31977776666", "PENDENTE", "",              "", ""),
    ("Carlos Mendes",  "11133334444", "PENDENTE", "Gamma S/A",     "", ""),
    ("Teste Invalido", "000",         "PENDENTE", "",              "", ""),
]

alt_fill = PatternFill("solid", fgColor="F0FAF7")
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if r % 2 == 0:
            cell.fill = alt_fill

widths = [24, 16, 14, 22, 28, 18]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

output = "contatos_modelo.xlsx"
wb.save(output)
print(f"Planilha criada: {output}")
print("Edite os dados de exemplo com seus contatos reais e defina Status=PENDENTE.")
