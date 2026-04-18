"""
Script para gerar o modelo de planilha de contatos.
Uso: python scripts/gerar_modelo_planilha.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def gerar_modelo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatos"
    
    # Cabeçalhos
    headers = ['nome', 'numero', 'empresa', 'adicional1', 'adicional2', 'adicional3']
    header_fill = PatternFill("solid", fgColor="2D7FF9")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Dados de exemplo
    exemplos = [
        ["João Silva",    "11999998888", "Silva LTDA",  "Produto X", "R$ 150,00", "30/04/2026"],
        ["Maria Santos",  "11988887777", "Santos ME",   "Produto Y", "R$ 280,00", "05/05/2026"],
        ["Carlos Lima",   "11977776666", "Lima & Cia",  "Produto Z", "R$ 99,00",  "10/05/2026"],
    ]
    
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    
    for row_idx, row_data in enumerate(exemplos, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.font = Font(name="Arial", size=11)
    
    # Ajustar largura das colunas
    col_widths = [20, 18, 20, 20, 15, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width
    
    import os
    # Garante que salva na raiz (subindo um nivel se estiver dentro de scripts ou apenas na raiz se cwd for SigmaLeeds)
    wb.save("contatos_modelo.xlsx")
    print("✓ Arquivo 'contatos_modelo.xlsx' criado na raiz do projeto.")
    print("\nVariáveis disponíveis para usar na mensagem:")
    for h in headers:
        print(f"  {{{h}}}")

if __name__ == "__main__":
    gerar_modelo()
